# -*- coding: utf-8 -*-
"""Gera o tracker do projeto (.xlsx) a partir do plano.json. O arquivo sobe para
o Drive e vira Google Sheets — as abas de visão usam FILTER, que o Sheets entende."""
import json, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

P = json.load(open("plano.json"))
TAR = P["tarefas"]; FAS = P["fases"]

TINTA   = "1B2A41"; AZUL = "2B4C8C"; CINZA = "EDF0F5"; BORDA = "C8D0DC"
STATUS  = ["Não iniciado","Em andamento","Bloqueado","Em revisão","Concluído"]
CORES_S = {"Não iniciado":"E8EBF0","Em andamento":"D6E4FF","Bloqueado":"FFD9D6",
           "Em revisão":"FFF0C9","Concluído":"D5F0DE"}

wb = Workbook(); wb.remove(wb.active)
fina = Side(style="thin", color=BORDA)
box  = Border(left=fina, right=fina, top=fina, bottom=fina)

def cabecalho(ws, cols, larguras, titulo=None):
    r = 1
    if titulo:
        ws.cell(1,1,titulo).font = Font(bold=True, size=14, color=TINTA); r = 3
    for j,(c,w) in enumerate(zip(cols,larguras), start=1):
        cel = ws.cell(r,j,c)
        cel.font = Font(bold=True, color="FFFFFF"); cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(vertical="center", wrap_text=True); cel.border = box
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[r].height = 30
    return r

def escrever(ws, r0, linhas, wrap_cols=()):
    for i, linha in enumerate(linhas, start=r0+1):
        for j, v in enumerate(linha, start=1):
            cel = ws.cell(i, j, v); cel.border = box
            cel.alignment = Alignment(vertical="top", wrap_text=(j in wrap_cols))
    return r0 + len(linhas)

# ---------- Tarefas: a fonte de verdade ----------
ws = wb.create_sheet("Tarefas")
cols = ["ID","Fase","Tarefa","Dono","Prioridade","Esforço (h)","Data alvo","Dependência",
        "Status","Bloqueio — motivo","Última atualização","Áreas envolvidas"]
larg = [8,7,62,9,11,12,12,16,15,34,17,22]
r0 = cabecalho(ws, cols, larg)
linhas = [[t["id"], t["fase"], t["titulo"], t["dono"], t["prioridade"], t["esforco_h"],
           dt.date.fromisoformat(t["data_alvo"]), t["dependencia"], t["status"], "", "", t["areas"]] for t in TAR]
fim = escrever(ws, r0, linhas, wrap_cols=(3,10,12))
for i in range(r0+1, fim+1):
    ws.cell(i,7).number_format = "dd/mm/yyyy"
    ws.cell(i,11).number_format = "dd/mm/yyyy"
ws.freeze_panes = ws.cell(r0+1,4)
ws.auto_filter.ref = f"A{r0}:L{fim}"

dv_s = DataValidation(type="list", formula1='"'+",".join(STATUS)+'"', allow_blank=False)
dv_p = DataValidation(type="list", formula1='"P0,P1,P2"', allow_blank=False)
dv_d = DataValidation(type="list", formula1='"Luana,Luiza"', allow_blank=False)
for dv, col in ((dv_s,"I"), (dv_p,"E"), (dv_d,"D")):
    ws.add_data_validation(dv); dv.add(f"{col}{r0+1}:{col}{fim}")
for st, cor in CORES_S.items():
    ws.conditional_formatting.add(f"A{r0+1}:L{fim}",
        FormulaRule(formula=[f'$I{r0+1}="{st}"'], fill=PatternFill("solid", fgColor=cor), stopIfTrue=False))
ws.conditional_formatting.add(f"G{r0+1}:G{fim}",
    FormulaRule(formula=[f'AND($G{r0+1}<TODAY(),$I{r0+1}<>"Concluído")'],
                font=Font(bold=True, color="B3261E"), stopIfTrue=False))

# ---------- Fases ----------
ws = wb.create_sheet("Fases")
cols = ["Fase","Nome","Dono","Apoio","Início","Fim","Objetivo","Entregável concreto","Critério de saída","Depende de"]
r0 = cabecalho(ws, cols, [7,30,9,9,12,12,52,46,60,12], "Fases do projeto — uma fase, um dono")
linhas = [[f["id"], f["nome"], f["dono"], f["apoio"], dt.date.fromisoformat(f["inicio"]),
           dt.date.fromisoformat(f["fim"]), f["objetivo"], f["entregavel"], f["saida"], f["depende"]] for f in FAS]
fim = escrever(ws, r0, linhas, wrap_cols=(2,7,8,9))
for i in range(r0+1, fim+1):
    for c in (5,6): ws.cell(i,c).number_format = "dd/mm/yyyy"
    ws.row_dimensions[i].height = 72

# ---------- Escopo ----------
ws = wb.create_sheet("Escopo")
ws["A1"] = "Escopo fechado"; ws["A1"].font = Font(bold=True, size=14, color=TINTA)
dentro = [
 "Desenhar e escolher o modelo de terceirização da operação do long tail regional",
 "Validar juridicamente o modelo e o formato de remuneração",
 "Definir a remuneração, as regras de apuração e o business case",
 "Redigir a minuta de contrato e a matriz de alçadas do representante",
 "Montar o onboarding do representante e o do varejo, tocado por ele",
 "Especificar e publicar o dash de operação e o dash financeiro com export CSV",
 "Publicar repositório de trade, brand book, FAQ e canais de atendimento",
 "Selecionar e contratar UM parceiro para o piloto",
 "Rodar o piloto por 3 ciclos de apuração e recomendar escalar, ajustar ou encerrar",
]
fora = [
 "Terceirizar varejo corporate — segue com o time interno da UME",
 "Prospecção e venda de novos varejos pelo representante (decisão Q1; fora do piloto de qualquer forma)",
 "Rebate e MDR para o representante — só existe se a Q1 virar sim",
 "Automatizar o cadastro de lojas e operadores — fica com o time da Amanda e vira backlog de Tech",
 "Modelo de franquia e licenciamento de marca — descartado em 26/ago",
 "Criar um time interno de N2 dedicado ao canal terceirizado",
 "Mudar política de crédito, alçadas de aprovação ou qualquer coisa em produção",
 "Onda 2 de representantes e expansão nacional — só depois do Gate 5",
 "Uniforme ou identidade visual obrigatória para o representante — vira recomendação, por risco de vínculo",
]
r = 3
ws.cell(r,1,"DENTRO do escopo").font = Font(bold=True, color="0E6B3F")
ws.cell(r,3,"FORA do escopo — não faremos agora").font = Font(bold=True, color="B3261E")
for i in range(max(len(dentro), len(fora))):
    if i < len(dentro): ws.cell(r+1+i,1,"• "+dentro[i]).alignment = Alignment(wrap_text=True, vertical="top")
    if i < len(fora):   ws.cell(r+1+i,3,"• "+fora[i]).alignment  = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 78; ws.column_dimensions["B"].width = 3; ws.column_dimensions["C"].width = 78
r = r + max(len(dentro), len(fora)) + 3
ws.cell(r,1,"Definição de pronto — o projeto acabou quando:").font = Font(bold=True, size=12, color=TINTA)
pronto = [
 "existe um representante terceirizado rodando uma carteira long tail com contrato assinado;",
 "a remuneração foi apurada e paga por 3 ciclos, com o business case real medido contra o modelado;",
 "o playbook (onboarding, guias, dashboards, FAQ) está publicado e foi usado por alguém de fora da UME;",
 "o time interno consegue mostrar, com número, quantas horas por semana deixou de gastar com esses varejos;",
 "a recomendação de escalar, ajustar ou encerrar foi entregue ao Lucas e decidida.",
]
for i,l in enumerate(pronto): ws.cell(r+1+i,1,"• "+l).alignment = Alignment(wrap_text=True)

# ---------- Métricas ----------
ws = wb.create_sheet("Métricas")
r0 = cabecalho(ws, ["ID","Métrica de sucesso","Baseline hoje","Meta","Dono","Onde se mede","Valor atual","Atualizado em"],
               [7,52,40,40,9,26,13,15], "Métricas de sucesso")
fim = escrever(ws, r0, [[m["id"], m["metrica"], m["baseline"], m["meta"], m["dono"], m["fonte"], "", ""] for m in P["metricas"]],
               wrap_cols=(2,3,4,6))
for i in range(r0+1, fim+1):
    ws.row_dimensions[i].height = 44; ws.cell(i,8).number_format = "dd/mm/yyyy"

# ---------- Riscos ----------
ws = wb.create_sheet("Riscos")
r0 = cabecalho(ws, ["ID","Risco","Probabilidade","Impacto","Sinal de alerta antecipado","Mitigação","Dono do risco","Status"],
               [7,54,14,10,50,58,13,14], "Top 5 riscos")
fim = escrever(ws, r0, [[r["id"], r["risco"], r["probabilidade"], r["impacto"], r["sinal"], r["mitigacao"], r["dono"], "Vigiando"]
                        for r in P["riscos"]], wrap_cols=(2,5,6))
for i in range(r0+1, fim+1): ws.row_dimensions[i].height = 64
ws.conditional_formatting.add(f"C{r0+1}:D{fim}",
    CellIsRule(operator="equal", formula=['"Alta"'], fill=PatternFill("solid", fgColor="FFD9D6")))
ws.conditional_formatting.add(f"D{r0+1}:D{fim}",
    CellIsRule(operator="equal", formula=['"Alto"'], fill=PatternFill("solid", fgColor="FFD9D6")))

# ---------- Carga semanal ----------
ws = wb.create_sheet("Carga semanal")
r0 = cabecalho(ws, ["Semana (segunda)","Capacidade (h)","Luana (h)","Luiza (h)","Folga Luana","Folga Luiza"],
               [18,16,12,12,13,13], "Realidade da capacidade — 8h/semana por pessoa, já descontado feriado")
linhas = [[dt.date.fromisoformat(c["semana"]), c["capacidade"], c["Luana"], c["Luiza"],
           round(c["capacidade"]-c["Luana"],1), round(c["capacidade"]-c["Luiza"],1)] for c in P["carga_semanal"]]
fim = escrever(ws, r0, linhas)
for i in range(r0+1, fim+1): ws.cell(i,1).number_format = "dd/mm/yyyy"
ws.conditional_formatting.add(f"E{r0+1}:F{fim}",
    CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFD9D6")))
ws.conditional_formatting.add(f"E{r0+1}:F{fim}",
    CellIsRule(operator="greaterThan", formula=["2"], fill=PatternFill("solid", fgColor="D5F0DE")))

# ---------- Views ----------
def view(nome, titulo, formula, nota):
    w = wb.create_sheet(nome)
    w["A1"] = titulo; w["A1"].font = Font(bold=True, size=14, color=TINTA)
    w["A2"] = nota;   w["A2"].font = Font(italic=True, size=10, color="55607A")
    r = cabecalho(w, ["ID","Fase","Tarefa","Dono","Prioridade","Esforço (h)","Data alvo","Dependência","Status","Bloqueio — motivo"],
                  [8,7,62,9,11,12,12,16,15,34])
    w.cell(r+1,1,formula)
    return w
ULT = len(TAR)+3
view("V · Vence em 7 dias", "O que vence nos próximos 7 dias",
     f'=SORT(FILTER(Tarefas!A4:J{ULT}, Tarefas!G4:G{ULT}<=TODAY()+7, Tarefas!I4:I{ULT}<>"Concluído"), 7, TRUE)',
     "Atualiza sozinha. Se aparecer vazia, não há nada vencendo — ou tudo já está concluído.")
view("V · Bloqueios", "Só o que está bloqueado",
     f'=FILTER(Tarefas!A4:J{ULT}, Tarefas!I4:I{ULT}="Bloqueado")',
     "Bloqueio se registra na hora, não na reunião. Passou de 3 dias úteis aqui, sobe para o Lucas.")
view("V · Luana", "Tarefas da Luana em aberto",
     f'=SORT(FILTER(Tarefas!A4:J{ULT}, Tarefas!D4:D{ULT}="Luana", Tarefas!I4:I{ULT}<>"Concluído"), 7, TRUE)',
     "Ordenada por data alvo.")
view("V · Luiza", "Tarefas da Luiza em aberto",
     f'=SORT(FILTER(Tarefas!A4:J{ULT}, Tarefas!D4:D{ULT}="Luiza", Tarefas!I4:I{ULT}<>"Concluído"), 7, TRUE)',
     "Ordenada por data alvo.")

# ---------- Para o Painel ----------
ws = wb.create_sheet("Para o Painel")
ws["A1"] = "Linhas prontas para colar no Importar do Painel de Demandas"
ws["A1"].font = Font(bold=True, size=14, color=TINTA)
ws["A2"] = "Formato do painel: cluster | conta | responsável | AAAA-MM-DD | descrição. Copie só as linhas da quinzena."
ws["A2"].font = Font(italic=True, size=10, color="55607A")
ws["A4"] = "Linha"; ws["A4"].font = Font(bold=True, color="FFFFFF"); ws["A4"].fill = PatternFill("solid", fgColor=AZUL)
ws.column_dimensions["A"].width = 130
for i in range(len(TAR)):
    ws.cell(5+i, 1, f'="novos | Terceirização | "&Tarefas!D{4+i}&" | "&TEXT(Tarefas!G{4+i},"yyyy-mm-dd")&" | ["&Tarefas!A{4+i}&"] "&Tarefas!C{4+i}')

# ---------- Como usar ----------
ws = wb.create_sheet("Como usar", 0)
ws.column_dimensions["A"].width = 116
bloco = [
 ("t","Tracker · Modelo de Terceirização da Operação"),
 ("i","Fonte única de verdade do projeto. Gerado em 26/ago/2026 a partir da reunião Escalabilidade time."),
 ("",""),
 ("h","A regra que faz isso funcionar"),
 ("p","A aba Tarefas manda. Toda tarefa mexida no dia é atualizada antes de você sair — não na reunião, no dia."),
 ("p","Bloqueio se registra na hora, com motivo escrito na coluna Bloqueio. Bloqueio sem motivo escrito não conta como bloqueio."),
 ("p","Bloqueio parado há mais de 3 dias úteis sobe para o Lucas, com a decisão pedida por escrito."),
 ("p","O Painel de Demandas (localhost:4321) fica com as tarefas da quinzena, para dar baixa no dia a dia. Quando o painel e o tracker divergirem, o tracker vence."),
 ("",""),
 ("h","Status possíveis"),
 ("p","Não iniciado · Em andamento · Bloqueado · Em revisão · Concluído. A coluna já vem com lista fechada — não invente status novo."),
 ("p","Em revisão significa: o trabalho acabou e está com outra pessoa ou outra área. Continua sendo seu até voltar."),
 ("",""),
 ("h","Prioridade"),
 ("p","P0 — sem ela o projeto não existe. P1 — importante, mas o projeto sobrevive. P2 — melhora o resultado."),
 ("p","Quando a semana apertar, corta P2, depois P1. P0 nunca escorrega em silêncio: escorregou, avisa no touchpoint."),
 ("",""),
 ("h","As quatro views"),
 ("p","Vence em 7 dias · Bloqueios · Luana · Luiza. São fórmulas, atualizam sozinhas. Não edite nada dentro delas — edite na aba Tarefas."),
 ("p","Para ver por fase, use o filtro da coluna Fase na própria aba Tarefas."),
 ("",""),
 ("h","O que é humano e o que dá para automatizar"),
 ("p","Humano: status, motivo do bloqueio, data de última atualização, valor atual das métricas."),
 ("p","Automático: as views, as linhas da aba Para o Painel, o realce de data vencida e o resumo de segunda-feira."),
 ("",""),
 ("h","Resumo de segunda, 9h"),
 ("p","O resumo lê este arquivo e manda no canal do projeto: o que andou na semana, o que travou e há quantos dias, o que vence nos próximos 7 dias e o que precisa de decisão do Lucas."),
 ("p","Se ninguém atualizar o tracker, o resumo chega vazio — e isso também é informação."),
 ("",""),
 ("h","Cadência"),
 ("p","Touchpoint terça e quinta, 30 min, só este assunto. Bloco protegido sexta de manhã, 2h. Status quinzenal com o Lucas, 15 min. Comitê de áreas mensal, 1h."),
]
r = 1
for tipo, txt in bloco:
    c = ws.cell(r,1,txt)
    if tipo == "t": c.font = Font(bold=True, size=16, color=TINTA)
    elif tipo == "i": c.font = Font(italic=True, size=11, color="55607A")
    elif tipo == "h": c.font = Font(bold=True, size=12, color=AZUL)
    else: c.font = Font(size=11); c.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[r].height = 30
    r += 1

wb.save("tracker-terceirizacao.xlsx")
print("ok — abas:", wb.sheetnames)
